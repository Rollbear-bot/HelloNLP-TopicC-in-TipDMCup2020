# -*- coding: utf-8 -*-
# @Time: 2020/3/28 16:41
# @Author: Rollbear
# @Filename: label.py
# “标签”的定义

import openpyxl


class Label:
    """标签类"""
    def __init__(self, name: str):
        """构造方法"""
        self.sub_label = []  # 子标签
        self.label_name = name  # 标签名

    def __eq__(self, other):
        return self.label_name == other.label_name

    def __str__(self):
        return self.label_name


def read_labels(path: str):
    """
    读取所有标签及其分级结构（附件一的格式）
    :param path: excel表路径
    :return: 根标签，Label对象
    """
    # 定义标签
    root_label = Label("根标签")

    # 打开excel工作簿
    e1 = openpyxl.load_workbook(path, read_only=True)
    sheet_names = e1.sheetnames  # 获取所有工作表名
    ws = e1[sheet_names[0]]  # 获取第一张工作表

    # 开始抓取标签
    for r in range(2, ws.max_row + 1):  # 忽略第一行（表头）
        cur_node = root_label  # 指针定位到根标签

        for c in range(1, 4):  # 遍历三列（分别为一二三级标签）
            cur_label = Label(ws.cell(row=r, column=c).value)
            try:
                index = cur_node.sub_label.index(cur_label)
            except ValueError:  # 这个标签还没有保存
                cur_node.sub_label.append(cur_label)  # 将这个标签加入子表
                index = cur_node.sub_label.index(cur_label)

            cur_node = cur_node.sub_label[index]
    return root_label
