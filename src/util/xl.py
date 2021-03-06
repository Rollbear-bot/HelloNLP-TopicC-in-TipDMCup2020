# -*- coding: utf-8 -*-
# @Time: 2020/3/28 17:19
# @Author: Rollbear
# @Filename: xl.py
# excel表格读取工具

import openpyxl  # 读excel
from pandas import DataFrame  # 写excel

from entity.label import LabelNode
from .timer import timer


@timer  # 如果不想显示运行时间，注释掉装饰器即可
def read_labels(path: str):
    """
    读取所有标签及其分级结构（附件一的格式）
    :param path: excel表路径
    :return: 根标签，Label对象
    """
    # 定义标签
    root_label = LabelNode("根标签")

    # 打开excel工作簿
    e1 = openpyxl.load_workbook(path, read_only=True)
    sheet_names = e1.sheetnames  # 获取所有工作表名
    ws = e1[sheet_names[0]]  # 获取第一张工作表

    # 开始抓取标签
    for r in range(2, ws.max_row + 1):  # 忽略第一行（表头）
        cur_node = root_label  # 指针定位到根标签

        for c in range(1, 4):  # 遍历三列（分别为一二三级标签）
            cur_label = LabelNode(ws.cell(row=r, column=c).value)
            try:
                index = cur_node.sub_label.index(cur_label)
            except ValueError:  # 这个标签还没有保存
                cur_node.sub_label.append(cur_label)  # 将这个标签加入子表
                index = cur_node.sub_label.index(cur_label)

            cur_node = cur_node.sub_label[index]
    return root_label


@timer
def read_labels_by_lines(path: str):
    """读取标签（附件一），返回列表"""
    rows = []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    row = ws[2:ws.max_row]  # 跳过第一行（表头）
    for r in row:
        row_tuple = tuple(map(lambda x: x.value, r))
        rows.append(row_tuple)
    return rows


@timer
def read_xl_by_line(path: str, skip_first_line=True):
    """
    读取表（通用）
    :param path: excel文件路径
    :param skip_first_line: 是否跳过首行
    :return: 列表，元素为n元组
    """
    res_rows = []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    if skip_first_line:
        row = ws[2:ws.max_row]  # 跳过第一行（表头）
    else:
        row = ws[1:ws.max_row]

    for r in row:
        # 将openpyxl的内置对象转化为元组
        row_tuple = tuple(map(lambda x: x.value, r))
        if row_tuple[0] is None:  # 到达空行时退出
            break
        res_rows.append(row_tuple)
    return res_rows


def write_rows(path: str, rows: list, title: tuple):
    """
    将元组写入Excel表格
    :param path: Excel表路径
    :param rows: 元组列表
    :param title: 表头
    """
    data = {key: [row[index] for row in rows] for index, key in enumerate(title)}

    # data生成的格式为：
    # data = {
    #    'name': ['张三', '李四', '王五'],
    #    'age': [11, 12, 13],
    #    'sex': ['男', '女', '男']
    # }

    df = DataFrame(data)
    df.to_excel(path, index=False)  # 存储到表格
