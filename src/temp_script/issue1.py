# -*- coding: utf-8 -*-
# @Time: 2020/5/8 9:44
# @Author: Rollbear
# @Filename: issue1.py
# 问题一：群众留言分类

import gensim
import openpyxl
from sklearn.externals import joblib  # 模型保存与加载

from util.dataset import *
from util.vec import doc_vec


def main():
    # 提示：待预测的表的输入路径在src/util/path.py中设置
    # 停用词表
    print("loading stop words...")
    stop = fetch_default_stop_words()

    # 从预设的路径加载knn分类器
    print("loading knn model...")
    knn = joblib.load(knn_model_path)

    # 从预设的路径加载词向量化模型Word2Vec
    print("loading word2vec model...")
    wv_model = gensim.models.KeyedVectors.load_word2vec_format(word2vec_model_path, binary=False)

    # 从预设的路径加载与kNN模型对应的标签名
    print("loading target names...")
    target_names = fetch_knn_target_names()

    # 预处理，并构造留言对象
    input_sheet = read_xl_by_line(predict_sheet_input_path)
    print(f"loaded {len(input_sheet)} records from {predict_sheet_input_path}.")
    print("preprocessing...")
    comm_dict = Comm.generate_comm_dict(input_sheet,
                                        stop_words_lt=stop,
                                        cut_all=True,
                                        full_dataset=True)
    # 计算文档向量，向量化模型使用预训练的Word2Vec
    vec = [doc_vec(comm_dict[row[0]].seg_detail, model=wv_model) for row in input_sheet]

    # 使用分类器预测分类
    print("predicting...")
    predicted = knn.predict(vec)

    # 将标签号转化为标签名
    predicted_labels = [target_names[index] for index in predicted]
    print("labeled.")

    # 至此，标签预测已经完成，预测的标签按照输入表的顺序存储在上面的predicted_labels列表中
    print(predicted_labels)

    # 将预测的信息写入到结果表中
    wb = openpyxl.load_workbook("../C_test/附件2（测试结果）.xlsx", read_only=False)
    ws = wb[wb.sheetnames[0]]  # 获取第一个sheet
    count = 2
    for label in predicted_labels:
        ws["B" + str(count)] = label
        count += 1
    wb.save("../附件2（测试结果）.xlsx")


if __name__ == '__main__':
    main()
